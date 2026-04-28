import re
import html
from collections import defaultdict

# ─────────────────────────────────────────────
# 1.  Categorías de caracteres (ajusta si necesitas)
# ─────────────────────────────────────────────
CATEGORIES = {
    "letter":    set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"),
    "digit":     set("0123456789"),
    "underscore":set("_"),
    "plus":      set("+"),
    "minus":     set("-"),
    "mult":      set("*"),
    "div":       set("/"),
    "lt":        set("<"),
    "gt":        set(">"),
    "eq":        set("="),
    "excl":      set("!"),
    "dquote":    set('"'),
    "dot":       set("."),
    "semicolon": set(";"),
    "comma":     set(","),
    "lparen":    set("("),
    "rparen":    set(")"),
    "lbracket":  set("["),
    "rbracket":  set("]"),
    "lcurly":    set("{"),
    "rcurly":    set("}"),
    "space":     set(" "),
    "tab":       set("\t"),
    "newline":   set("\n"),
}

# Nombres bonitos para la tabla del reporte
CATEGORY_NAMES = {
    "letter":    "letter",
    "digit":     "digit",
    "underscore":"_",
    "plus":      "+",
    "minus":     "-",
    "mult":      "*",
    "div":       "/",
    "lt":        "<",
    "gt":        ">",
    "eq":        "=",
    "excl":      "!",
    "dquote":    '"',
    "dot":       ".",
    "semicolon": ";",
    "comma":     ",",
    "lparen":    "(",
    "rparen":    ")",
    "lbracket":  "[",
    "rbracket":  "]",
    "lcurly":    "{",
    "rcurly":    "}",
    "space":     "space",
    "tab":       "tab",
    "newline":   "\\n",
}

# Token IDs para nombres legibles en la tabla
TOKEN_NAMES = {
    "1":  "KW_INT",
    "2":  "KW_FLOAT",
    "3":  "KW_STRING",
    "4":  "KW_RETURN",
    "5":  "KW_IF",
    "6":  "KW_ELSE",
    "7":  "KW_WHILE",
    "8":  "KW_READ",
    "9":  "KW_WRITE",
    "10": "KW_VOID",
    "11": "TK_PLUS",
    "12": "TK_MINUS",
    "13": "TK_MULT",
    "14": "TK_DIV",
    "15": "TK_LT",
    "16": "TK_LEQ",
    "17": "TK_GT",
    "18": "TK_GEQ",
    "19": "TK_EQ",
    "20": "TK_NEQ",
    "21": "TK_ASSIGN",
    "22": "TK_SEMI",
    "23": "TK_COMMA",
    "24": "TK_DQUOTE",
    "25": "TK_LPAREN",
    "26": "TK_RPAREN",
    "27": "TK_LBRACKET",
    "28": "TK_RBRACKET",
    "29": "TK_LCURLY",
    "30": "TK_RCURLY",
    "31": "TK_ID",
    "32": "TK_INT_NUM",
    "33": "TK_FLOAT_NUM",
    "34": "TK_STRING_LIT",
    "35": "ERROR",
    "36": "COMMENT",
}

# ─────────────────────────────────────────────
# 2.  Parsear el .jff
# ─────────────────────────────────────────────
def extract_data(file_location):
    with open(file_location, 'r') as f:
        xml = f.read()
    state_data      = "\n".join(re.findall(r'<state id="\d+".*?</state>', xml, re.DOTALL))
    transition_data = "\n".join(re.findall(r'<transition>.*?</transition>', xml, re.DOTALL))
    return state_data, transition_data


def parse_states(state_data):
    states, final_states, states_label = [], [], {}
    initial_state = None
    for m in re.finditer(r'<state id="(\d+)".*?</state>', state_data, re.DOTALL):
        sid   = int(m.group(1))
        label = re.search(r'<label>(\d+)</label>', m.group(0))
        states.append(sid)
        states_label[sid] = label.group(1) if label else None
        if re.search(r'<initial/>', m.group(0)):
            initial_state = sid
        if re.search(r'<final/>',   m.group(0)):
            final_states.append(sid)
    return states, initial_state, final_states, states_label


def parse_transitions(transition_data):
    transitions = []
    for m in re.finditer(r'<transition>(.*?)</transition>', transition_data, re.DOTALL):
        frm  = re.search(r'<from>(\d+)</from>',   m.group(1)).group(1)
        to   = re.search(r'<to>(\d+)</to>',       m.group(1)).group(1)
        read = re.search(r'<read>(.*?)</read>',    m.group(1)).group(1)
        transitions.append((int(frm), int(to), html.unescape(read)))
    return transitions

# ─────────────────────────────────────────────
# 3.  Asignar cada etiqueta JFLAP a una categoría
# ─────────────────────────────────────────────
def char_to_category(ch):
    """Devuelve la categoría de un carácter o etiqueta de JFLAP."""
    for cat, chars in CATEGORIES.items():
        if ch in chars:
            return cat
    return "other"   # carácter no reconocido → columna "other"


def label_to_categories(label):
    """
    Convierte una etiqueta de JFLAP (puede ser un char real o un nombre
    como 'diagonal', 'asterisco', etc.) en una lista de categorías.
    """
    # Mapeo de etiquetas textuales que usaste en JFLAP
    TEXT_MAP = {
        "diagonal":                     ["div"],
        "asterisco":                    ["mult"],
        "letter":                       ["letter"],
        "number":                       ["digit"],
        "underscore":                   ["underscore"],
        "letter|number|underscore":     ["letter", "digit", "underscore"],
        "letter|number|underscore ":    ["letter", "digit", "underscore"],
        "number|plus|minus":            ["digit", "plus", "minus"],
        "todos menos asterisco":        None,   # especial
        "todo menos comillas y new line": None, # especial
        "nothing|other than #":         None,
        "nothing":                      [],
        "comillas dobles":              ["dquote"],
        "space|tab|newLine":            ["space", "tab", "newline"],
        "new Line":                     ["newline"],
        "end of file":                  [],     # EOF no es un char
        "invalid character":            ["other"],
        "otro":                         ["other"],
        "open parenthesis":             ["lparen"],
        "close parenthesis":            ["rparen"],
        "open curly bracket":           ["lcurly"],
        "close curly bracket":          ["rcurly"],
        "open square bracket":          ["lbracket"],
        "close square bracket":         ["rbracket"],
        "semicolon":                    ["semicolon"],
        "coma":                         ["comma"],
        "sum":                          ["plus"],
        "minus":                        ["minus"],
        "multiplication":               ["mult"],
    }

    if label in TEXT_MAP:
        result = TEXT_MAP[label]
        return result if result is not None else ["other"]

    # Si es un carácter real de un solo símbolo
    if len(label) == 1:
        return [char_to_category(label)]

    # Fallback
    return ["other"]

# ─────────────────────────────────────────────
# 4.  Construir la tabla agrupada
# ─────────────────────────────────────────────
def build_grouped_table(states, transitions, final_states, states_label):
    # Columnas que aparecerán en la tabla (en orden)
    COL_ORDER = [
        "letter","digit","underscore",
        "plus","minus","mult","div",
        "lt","gt","eq","excl",
        "dquote","dot",
        "semicolon","comma",
        "lparen","rparen","lbracket","rbracket","lcurly","rcurly",
        "space","tab","newline",
        "other",
    ]

    # tabla[estado][categoria] = estado_destino  (-1 = error)
    table = defaultdict(lambda: defaultdict(lambda: -1))

    for frm, to, label in transitions:
        cats = label_to_categories(label)
        for cat in cats:
            # Si ya hay un valor distinto, lo dejamos (primera regla gana)
            if table[frm][cat] == -1:
                table[frm][cat] = to

    return table, COL_ORDER


def cell_str(state, cat, table, final_states, states_label):
    """Devuelve el texto de una celda."""
    dst = table[state][cat]
    if dst == -1:
        return "ERR"
    # Si el destino es final, mostramos el nombre del token
    if dst in final_states:
        label = states_label.get(dst)
        if label and label in TOKEN_NAMES:
            return f"q{dst}\n({TOKEN_NAMES[label]})"
        return f"q{dst}(*)"
    return f"q{dst}"

# ─────────────────────────────────────────────
# 5.  Generar el reporte en TXT y en HTML
# ─────────────────────────────────────────────
def write_txt(states, table, col_order, final_states, states_label, filename):
    COL_W = 14

    header_cats = [CATEGORY_NAMES.get(c, c) for c in col_order]

    with open(filename, 'w', encoding='utf-8') as f:
        f.write("TRANSITION TABLE — C-- Lexical Analyzer\n")
        f.write("=" * 80 + "\n\n")

        # ── Token legend ──
        f.write("TOKEN LEGEND\n")
        f.write("-" * 40 + "\n")
        for k, v in TOKEN_NAMES.items():
            lbl = states_label.get(None)  # placeholder
            f.write(f"  {v:<20} = {k}\n")
        f.write("\n")

        # ── Final states ──
        f.write("FINAL / ACCEPTANCE STATES\n")
        f.write("-" * 40 + "\n")
        for s in sorted(final_states):
            lbl   = states_label.get(s)
            tname = TOKEN_NAMES.get(lbl, "?") if lbl else "—"
            f.write(f"  q{s:<4}  label={lbl or '—':<4}  token={tname}\n")
        f.write("\n")

        # ── Table header ──
        state_col = 8
        line = f"{'State':<{state_col}}"
        for h in header_cats:
            line += f"{h:^{COL_W}}"
        f.write(line + "\n")
        f.write("-" * (state_col + COL_W * len(col_order)) + "\n")

        # ── Rows ──
        for s in sorted(states):
            is_final = s in final_states
            lbl      = states_label.get(s)
            tname    = TOKEN_NAMES.get(lbl, "") if lbl else ""
            marker   = f"*{tname}" if is_final else ""
            row = f"{'q'+str(s)+marker:<{state_col}}"
            for cat in col_order:
                dst = table[s][cat]
                if dst == -1:
                    cell = "—"
                elif dst in final_states:
                    dl   = states_label.get(dst)
                    dn   = TOKEN_NAMES.get(dl, "") if dl else ""
                    cell = f"q{dst}" + (f"/{dn}" if dn else "")
                else:
                    cell = f"q{dst}"
                row += f"{cell:^{COL_W}}"
            f.write(row + "\n")

    print(f"[TXT] Written → {filename}")


def write_xlsx(states, table, col_order, final_states, states_label, filename):
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)

    wb = Workbook()

    # ── Sheet 1: Transition Table ──────────────────────────────────────
    ws = wb.active
    ws.title = "Transition Table"

    DARK_HEADER  = "1F3864"   # dark blue
    GREEN_ACCEPT = "C6EFCE"   # light green  (acceptance state row)
    GREEN_FONT   = "276221"
    YELLOW_CELL  = "FFEB9C"   # light yellow (transition → acceptance)
    YELLOW_FONT  = "9C6500"
    ERR_COLOR    = "F4CCCC"   # light red
    WHITE        = "FFFFFF"
    GRAY_ROW     = "F2F2F2"

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_cats = [CATEGORY_NAMES.get(c, c) for c in col_order]

    # Title row
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=1 + len(col_order))
    title_cell = ws.cell(row=1, column=1,
                         value="Transition Table — C-- Lexical Analyzer")
    title_cell.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
    title_cell.fill      = PatternFill("solid", fgColor=DARK_HEADER)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Column header row
    ws.cell(row=2, column=1, value="State").font = Font(bold=True, color=WHITE, name="Arial")
    ws.cell(row=2, column=1).fill = PatternFill("solid", fgColor=DARK_HEADER)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=1).border = border

    for ci, hdr in enumerate(header_cats, start=2):
        c = ws.cell(row=2, column=ci, value=hdr)
        c.font      = Font(bold=True, color=WHITE, name="Arial", size=10)
        c.fill      = PatternFill("solid", fgColor=DARK_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border    = border
    ws.row_dimensions[2].height = 32

    # Data rows
    for ri, s in enumerate(sorted(states), start=3):
        is_final = s in final_states
        lbl      = states_label.get(s)
        tname    = TOKEN_NAMES.get(lbl, "") if lbl else ""
        state_label = f"q{s}" + (f" ({tname})" if tname else "")

        row_fill = PatternFill("solid", fgColor=GREEN_ACCEPT if is_final
                               else (GRAY_ROW if ri % 2 == 0 else WHITE))

        sc = ws.cell(row=ri, column=1, value=state_label)
        sc.font      = Font(bold=is_final, name="Arial", size=10,
                            color=(GREEN_FONT if is_final else "000000"))
        sc.fill      = row_fill
        sc.alignment = Alignment(horizontal="center", vertical="center")
        sc.border    = border

        for ci, cat in enumerate(col_order, start=2):
            dst = table[s][cat]
            if dst == -1:
                val  = "—"
                fill = PatternFill("solid", fgColor=ERR_COLOR)
                fnt  = Font(name="Arial", size=10, color="CC0000")
            elif dst in final_states:
                dl   = states_label.get(dst)
                dn   = TOKEN_NAMES.get(dl, "") if dl else ""
                val  = f"q{dst}" + (f" ({dn})" if dn else "")
                fill = PatternFill("solid", fgColor=YELLOW_CELL)
                fnt  = Font(name="Arial", size=10, color=YELLOW_FONT,
                            bold=True)
            else:
                val  = f"q{dst}"
                fill = row_fill
                fnt  = Font(name="Arial", size=10)

            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = fnt
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

    # Column widths
    ws.column_dimensions["A"].width = 18
    for ci in range(2, 2 + len(col_order)):
        col_letter = ws.cell(row=2, column=ci).column_letter
        ws.column_dimensions[col_letter].width = 13
    ws.freeze_panes = "B3"

    # ── Sheet 2: Token Legend ──────────────────────────────────────────
    wl = wb.create_sheet("Token Legend")
    wl.column_dimensions["A"].width = 8
    wl.column_dimensions["B"].width = 20
    wl.column_dimensions["C"].width = 14

    for col, hdr in enumerate(["ID", "Token Name", "Description"], start=1):
        c = wl.cell(row=1, column=col, value=hdr)
        c.font      = Font(bold=True, color=WHITE, name="Arial")
        c.fill      = PatternFill("solid", fgColor=DARK_HEADER)
        c.alignment = Alignment(horizontal="center")
        c.border    = border

    DESCRIPTIONS = {
        "1":"Keyword int",        "2":"Keyword float",
        "3":"Keyword string",     "4":"Keyword return",
        "5":"Keyword if",         "6":"Keyword else",
        "7":"Keyword while",      "8":"Keyword read",
        "9":"Keyword write",      "10":"Keyword void",
        "11":"Operator +",        "12":"Operator -",
        "13":"Operator *",        "14":"Operator /",
        "15":"Operator <",        "16":"Operator <=",
        "17":"Operator >",        "18":"Operator >=",
        "19":"Operator ==",       "20":"Operator !=",
        "21":"Assignment =",      "22":"Semicolon ;",
        "23":"Comma ,",           "24":"Double quote \"",
        "25":"Open parenthesis",  "26":"Close parenthesis",
        "27":"Open bracket [",    "28":"Close bracket ]",
        "29":"Open curly {",      "30":"Close curly }",
        "31":"Identifier / ID",   "32":"Integer number",
        "33":"Float number",      "34":"String literal",
        "35":"Error",             "36":"Comment",
    }

    for ri, (k, v) in enumerate(TOKEN_NAMES.items(), start=2):
        bg = GRAY_ROW if ri % 2 == 0 else WHITE
        for ci, val in enumerate([k, v, DESCRIPTIONS.get(k, "")], start=1):
            c = wl.cell(row=ri, column=ci, value=val)
            c.font      = Font(name="Arial", size=10)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center" if ci < 3 else "left",
                                    vertical="center")
            c.border    = border

    # ── Sheet 3: Acceptance States ────────────────────────────────────
    wa = wb.create_sheet("Acceptance States")
    wa.column_dimensions["A"].width = 10
    wa.column_dimensions["B"].width = 10
    wa.column_dimensions["C"].width = 20

    for col, hdr in enumerate(["State", "Label", "Token"], start=1):
        c = wa.cell(row=1, column=col, value=hdr)
        c.font      = Font(bold=True, color=WHITE, name="Arial")
        c.fill      = PatternFill("solid", fgColor=DARK_HEADER)
        c.alignment = Alignment(horizontal="center")
        c.border    = border

    for ri, s in enumerate(sorted(final_states), start=2):
        lbl   = states_label.get(s)
        tname = TOKEN_NAMES.get(lbl, "—") if lbl else "—"
        for ci, val in enumerate([f"q{s}", lbl or "—", tname], start=1):
            c = wa.cell(row=ri, column=ci, value=val)
            c.font      = Font(name="Arial", size=10)
            c.fill      = PatternFill("solid", fgColor=GREEN_ACCEPT)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = border

    wb.save(filename)
    print(f"[XLSX] Written → {filename}")


def write_html(states, table, col_order, final_states, states_label, filename):
    header_cats = [CATEGORY_NAMES.get(c, c) for c in col_order]

    rows_html = ""
    for s in sorted(states):
        is_final = s in final_states
        lbl      = states_label.get(s)
        tname    = TOKEN_NAMES.get(lbl, "") if lbl else ""
        style    = ' style="background:#d4edda;font-weight:bold"' if is_final else ""
        state_label = f"q{s}" + (f"<br><small>({tname})</small>" if tname else "")

        cells = f"<td{style}>{state_label}</td>"
        for cat in col_order:
            dst = table[s][cat]
            if dst == -1:
                cells += '<td style="color:#aaa">—</td>'
            elif dst in final_states:
                dl    = states_label.get(dst)
                dn    = TOKEN_NAMES.get(dl, "") if dl else ""
                inner = f"q{dst}" + (f"<br><small>({dn})</small>" if dn else "")
                cells += f'<td style="background:#fff3cd">{inner}</td>'
            else:
                cells += f"<td>q{dst}</td>"
        rows_html += f"<tr>{cells}</tr>\n"

    headers = "<th>State</th>" + "".join(f"<th>{h}</th>" for h in header_cats)

    html_out = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Transition Table — C--</title>
<style>
  body  {{ font-family: monospace; font-size: 12px; padding: 20px; }}
  table {{ border-collapse: collapse; white-space: nowrap; }}
  th,td {{ border: 1px solid #999; padding: 4px 8px; text-align: center; }}
  th    {{ background: #343a40; color: #fff; }}
  h1    {{ font-size: 18px; }}
</style>
</head>
<body>
<h1>Transition Table — C-- Lexical Analyzer</h1>
<p>
  <span style="background:#d4edda;padding:2px 6px">Green row</span> = acceptance state &nbsp;
  <span style="background:#fff3cd;padding:2px 6px">Yellow cell</span> = transition to acceptance state
</p>
<div style="overflow-x:auto">
<table>
<thead><tr>{headers}</tr></thead>
<tbody>
{rows_html}
</tbody>
</table>
</div>
</body>
</html>"""

    with open(filename, 'w', encoding='utf-8') as f:
        f.write(html_out)
    print(f"[HTML] Written → {filename}")


# ─────────────────────────────────────────────
# 6.  Main
# ─────────────────────────────────────────────
if __name__ == "__main__":
    JFF_FILE = "./FullAutomaton.jff"

    state_data, transition_data = extract_data(JFF_FILE)
    states, initial_state, final_states, states_label = parse_states(state_data)
    transitions = parse_transitions(transition_data)

    table, col_order = build_grouped_table(states, transitions, final_states, states_label)

    write_txt( states, table, col_order, final_states, states_label, "./transition_table2.txt")
    write_html(states, table, col_order, final_states, states_label, "./transition_table2.html")
    write_xlsx(states, table, col_order, final_states, states_label, "./transition_table2.xlsx")

    print(f"\nStates : {len(states)}")
    print(f"Finals : {len(final_states)}")
    print(f"Columns: {len(col_order)}")